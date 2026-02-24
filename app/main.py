"""
BINDMANAGER PRO — Backend FastAPI
"""
import os, json, io, logging
from datetime import datetime
from typing import Optional, List
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import engine, get_db, SessionLocal
from app import models
from app.auth import (hash_password, verify_password, create_token,
                      get_current_user, require_admin)
from app.import_excel import row_to_prodotto, valida_riga
from app.categorie_mapping import mappa_categoria
from app.ai_claude import genera_descrizione_claude, genera_descrizione_libro

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── APP ─────────────────────────────────────────────────────────
app = FastAPI(title="BindManager Pro", version="1.0.0")

app.add_middleware(CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"])

# ── DB INIT ─────────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    models.Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        # Crea admin di default se non esiste
        admin = db.query(models.User).filter(models.User.ruolo == "admin").first()
        if not admin:
            admin = models.User(
                email="admin@bindmanager.it",
                nome="Daniele",
                hashed_password=hash_password("bindmanager2024"),
                ruolo="admin"
            )
            db.add(admin)
            db.commit()
            logger.info("Admin creato: admin@bindmanager.it / bindmanager2024")
    finally:
        db.close()

# ── STATIC FILES ────────────────────────────────────────────────
static_dir = os.path.join(os.path.dirname(__file__), "..", "static")
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# ── PYDANTIC MODELS ─────────────────────────────────────────────
class UserCreate(BaseModel):
    email: str
    nome: str
    password: str
    ruolo: str = "collaboratore"

class ProdottoUpdate(BaseModel):
    titolo: Optional[str]
    descrizione: Optional[str]
    prezzo_vendita: Optional[float]
    quantita: Optional[int]
    stato_bene: Optional[str]
    tag: Optional[str]

class ConfigSet(BaseModel):
    chiave: str
    valore: str

# ── HELPER ──────────────────────────────────────────────────────
def _get_config(db: Session, chiave: str, default="") -> str:
    c = db.query(models.Configurazione).filter(models.Configurazione.chiave == chiave).first()
    return c.valore if c else default

def _set_config(db: Session, chiave: str, valore: str):
    c = db.query(models.Configurazione).filter(models.Configurazione.chiave == chiave).first()
    if c:
        c.valore = valore
    else:
        c = models.Configurazione(chiave=chiave, valore=valore)
        db.add(c)
    db.commit()

def _log(db: Session, operazione: str, dettaglio: str, livello="info", utente=""):
    db.add(models.Log(operazione=operazione, dettaglio=dettaglio,
                      livello=livello, utente_email=utente))
    db.commit()

# ── AUTH ────────────────────────────────────────────────────────
@app.post("/api/auth/login")
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form.username).first()
    if not user or not verify_password(form.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    if not user.attivo:
        raise HTTPException(status_code=403, detail="Account disattivato")
    token = create_token({"sub": user.email, "ruolo": user.ruolo})
    return {"access_token": token, "token_type": "bearer",
            "nome": user.nome, "ruolo": user.ruolo, "email": user.email}

@app.get("/api/auth/me")
def me(current_user=Depends(get_current_user)):
    return {"email": current_user.email, "nome": current_user.nome, "ruolo": current_user.ruolo}

# ── DASHBOARD ───────────────────────────────────────────────────
@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    totale = db.query(models.Prodotto).count()
    esauriti = db.query(models.Prodotto).filter(models.Prodotto.quantita == 0).count()
    sottoscorta = db.query(models.Prodotto).filter(
        models.Prodotto.quantita > 0, models.Prodotto.quantita <= 2).count()
    con_ai = db.query(models.Prodotto).filter(models.Prodotto.ai_generata == True).count()
    inviati_bc = db.query(models.Prodotto).filter(models.Prodotto.bindcommerce_inviato == True).count()
    con_foto = db.query(models.Prodotto).filter(models.Prodotto.foto1 != None,
                                                 models.Prodotto.foto1 != "").count()
    logs = db.query(models.Log).order_by(models.Log.creato_il.desc()).limit(20).all()

    return {
        "totale": totale,
        "esauriti": esauriti,
        "sottoscorta": sottoscorta,
        "con_ai": con_ai,
        "inviati_bc": inviati_bc,
        "con_foto": con_foto,
        "logs": [{"operazione": l.operazione, "dettaglio": l.dettaglio,
                  "livello": l.livello, "data": l.creato_il.isoformat() if l.creato_il else ""}
                 for l in logs]
    }

# ── PRODOTTI ────────────────────────────────────────────────────
@app.get("/api/prodotti")
def lista_prodotti(
    page: int = 1, per_page: int = 50,
    cerca: str = "", categoria: str = "", stato: str = "",
    solo_esauriti: bool = False, solo_senza_ai: bool = False,
    db: Session = Depends(get_db), current_user=Depends(get_current_user)
):
    q = db.query(models.Prodotto)
    if cerca:
        like = f"%{cerca}%"
        q = q.filter(
            models.Prodotto.titolo.ilike(like) |
            models.Prodotto.codice.ilike(like) |
            models.Prodotto.ean.ilike(like) |
            models.Prodotto.autore.ilike(like)
        )
    if categoria:
        q = q.filter(models.Prodotto.categoria1.ilike(f"%{categoria}%"))
    if stato:
        q = q.filter(models.Prodotto.stato_bene == stato)
    if solo_esauriti:
        q = q.filter(models.Prodotto.quantita == 0)
    if solo_senza_ai:
        q = q.filter(models.Prodotto.ai_generata == False)

    totale = q.count()
    prodotti = q.order_by(models.Prodotto.id.desc()).offset((page-1)*per_page).limit(per_page).all()

    return {
        "totale": totale,
        "pagine": (totale + per_page - 1) // per_page,
        "pagina": page,
        "prodotti": [{
            "id": p.id, "codice": p.codice, "ean": p.ean,
            "titolo": p.titolo, "autore": p.autore, "anno": p.anno,
            "categoria1": p.categoria1, "categoria2": p.categoria2,
            "produttore": p.produttore, "stato_bene": p.stato_bene,
            "prezzo_vendita": p.prezzo_vendita, "quantita": p.quantita,
            "foto1": p.foto1, "ai_generata": p.ai_generata,
            "bindcommerce_inviato": p.bindcommerce_inviato,
            "posizione_magazzino": p.posizione_magazzino,
        } for p in prodotti]
    }

@app.get("/api/prodotti/{prodotto_id}")
def get_prodotto(prodotto_id: int, db: Session = Depends(get_db),
                 current_user=Depends(get_current_user)):
    p = db.query(models.Prodotto).filter(models.Prodotto.id == prodotto_id).first()
    if not p:
        raise HTTPException(404, "Prodotto non trovato")
    return {
        "id": p.id, "codice": p.codice, "ean": p.ean,
        "titolo": p.titolo, "autore": p.autore, "anno": p.anno,
        "categoria1": p.categoria1, "categoria2": p.categoria2,
        "categoria3": p.categoria3, "categoria4": p.categoria4,
        "tipo_prodotto": p.tipo_prodotto, "produttore": p.produttore,
        "stato_bene": p.stato_bene, "specifiche": p.specifiche,
        "dettaglio_stato": p.dettaglio_stato,
        "prezzo_vendita": p.prezzo_vendita, "quantita": p.quantita,
        "posizione_magazzino": p.posizione_magazzino,
        "descrizione_breve": p.descrizione_breve,
        "descrizione": p.descrizione, "descrizione_sito": p.descrizione_sito,
        "note": p.note, "tag": p.tag,
        "foto1": p.foto1, "foto2": p.foto2, "foto3": p.foto3,
        "ai_generata": p.ai_generata, "bindcommerce_inviato": p.bindcommerce_inviato,
    }

@app.patch("/api/prodotti/{prodotto_id}")
def aggiorna_prodotto(prodotto_id: int, data: ProdottoUpdate,
                      db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    p = db.query(models.Prodotto).filter(models.Prodotto.id == prodotto_id).first()
    if not p:
        raise HTTPException(404, "Prodotto non trovato")
    for k, v in data.dict(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    return {"ok": True}

# ── IMPORT EXCEL ────────────────────────────────────────────────
@app.post("/api/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    genera_ai: bool = Form(False),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl non installato")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    importati = 0
    aggiornati = 0
    errori = []
    skippati = 0

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    api_key = _get_config(db, "claude_api_key")

    for i, row in enumerate(rows):
        try:
            prodotto_dict = row_to_prodotto(row)
            if not prodotto_dict:
                skippati += 1
                continue

            warnings = valida_riga(prodotto_dict)

            # Mappa categorie verso BindCommerce
            bc_cat = mappa_categoria(prodotto_dict.get("categoria1",""),
                                      prodotto_dict.get("categoria2",""))

            # Controlla se esiste già per codice
            esistente = None
            if prodotto_dict.get("codice"):
                esistente = db.query(models.Prodotto).filter(
                    models.Prodotto.codice == prodotto_dict["codice"]).first()

            if esistente:
                # Aggiorna quantità e prezzo
                esistente.quantita = prodotto_dict["quantita"]
                esistente.prezzo_vendita = prodotto_dict["prezzo_vendita"]
                esistente.stato_bene = prodotto_dict["stato_bene"] or esistente.stato_bene
                if prodotto_dict.get("foto1") and not esistente.foto1:
                    esistente.foto1 = prodotto_dict["foto1"]
                db.commit()
                aggiornati += 1
            else:
                p = models.Prodotto(**prodotto_dict)
                db.add(p)
                db.flush()

                # Genera AI se richiesto e disponibile
                if genera_ai and api_key and not prodotto_dict.get("descrizione"):
                    cat1 = prodotto_dict.get("categoria1", "").lower()
                    titolo = prodotto_dict.get("titolo", "")
                    autore = prodotto_dict.get("autore", "")
                    anno = prodotto_dict.get("anno", "")
                    produttore = prodotto_dict.get("produttore", "")
                    stato = prodotto_dict.get("stato_bene", "")
                    specifiche = prodotto_dict.get("specifiche", "")

                    if cat1 in ("cd", "dischi vinile 33 giri", "dischi vinile 45 giri", "musicassette"):
                        ai_data = genera_descrizione_claude(
                            artista=autore, album=titolo, anno=anno,
                            label=produttore, generi=cat1, stili="",
                            tracklist="", paese="", stato_fisico=stato,
                            specifiche=specifiche, api_key=api_key
                        )
                    else:
                        ai_data = genera_descrizione_libro(
                            titolo=titolo, autore=autore, anno=anno,
                            editore=produttore, stato=stato, api_key=api_key
                        )

                    if ai_data:
                        p.descrizione = ai_data.get("ebay_description", "")
                        p.descrizione_sito = ai_data.get("sito_description", "")
                        tags_extra = ai_data.get("tags_extra", [])
                        if tags_extra:
                            existing_tags = p.tag or ""
                            p.tag = existing_tags + (", " if existing_tags else "") + ", ".join(tags_extra)
                        p.ai_generata = True

                db.commit()
                importati += 1

        except Exception as e:
            errori.append(f"Riga {i+2}: {str(e)[:80]}")
            if len(errori) > 20:
                break

    _log(db, "Import Excel", f"Importati: {importati}, Aggiornati: {aggiornati}, "
         f"Errori: {len(errori)}, Skippati: {skippati}", "success", current_user.email)

    return {
        "importati": importati,
        "aggiornati": aggiornati,
        "errori": errori,
        "skippati": skippati,
        "totale_righe": len(rows)
    }

# ── AI: GENERA/RIGENERA DESCRIZIONE ────────────────────────────
@app.post("/api/prodotti/{prodotto_id}/genera-ai")
def genera_ai_prodotto(prodotto_id: int, db: Session = Depends(get_db),
                       current_user=Depends(get_current_user)):
    p = db.query(models.Prodotto).filter(models.Prodotto.id == prodotto_id).first()
    if not p:
        raise HTTPException(404, "Prodotto non trovato")

    api_key = _get_config(db, "claude_api_key")
    if not api_key:
        raise HTTPException(400, "API Key Claude non configurata — vai in Impostazioni")

    cat1 = (p.categoria1 or "").lower()
    if cat1 in ("cd", "dischi vinile 33 giri", "dischi vinile 45 giri", "musicassette"):
        ai_data = genera_descrizione_claude(
            artista=p.autore or "", album=p.titolo or "", anno=p.anno or "",
            label=p.produttore or "", generi=cat1, stili="",
            tracklist="", paese="", stato_fisico=p.stato_bene or "",
            specifiche=p.specifiche or "", api_key=api_key
        )
    else:
        ai_data = genera_descrizione_libro(
            titolo=p.titolo or "", autore=p.autore or "", anno=p.anno or "",
            editore=p.produttore or "", stato=p.stato_bene or "", api_key=api_key
        )

    if not ai_data:
        raise HTTPException(500, "Claude non ha generato una risposta valida")

    p.descrizione = ai_data.get("ebay_description", "")
    p.descrizione_sito = ai_data.get("sito_description", "")
    tags_extra = ai_data.get("tags_extra", [])
    if tags_extra:
        p.tag = (p.tag or "") + (", " if p.tag else "") + ", ".join(tags_extra)
    p.ai_generata = True
    db.commit()

    _log(db, "AI Generata", f"Descrizione generata per: {p.codice} — {p.titolo[:50]}",
         "success", current_user.email)
    return {"ok": True, "descrizione": p.descrizione, "tag": p.tag}

# ── CONFIGURAZIONE ──────────────────────────────────────────────
@app.get("/api/config")
def get_config(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    keys = ["claude_api_key", "bindcommerce_api_key", "sync_interval",
            "backblaze_key_id", "backblaze_app_key", "backblaze_bucket"]
    result = {}
    for k in keys:
        v = _get_config(db, k)
        # Maschera API key per sicurezza
        if "api_key" in k and v:
            result[k] = v[:8] + "..." + v[-4:] if len(v) > 12 else "****"
        else:
            result[k] = v
    return result

@app.post("/api/config")
def set_config(data: ConfigSet, db: Session = Depends(get_db),
               current_user=Depends(require_admin)):
    _set_config(db, data.chiave, data.valore)
    _log(db, "Config aggiornata", f"Chiave: {data.chiave}", "info", current_user.email)
    return {"ok": True}

# ── UTENTI ──────────────────────────────────────────────────────
@app.get("/api/utenti")
def lista_utenti(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    utenti = db.query(models.User).all()
    return [{"id": u.id, "email": u.email, "nome": u.nome,
             "ruolo": u.ruolo, "attivo": u.attivo} for u in utenti]

@app.post("/api/utenti")
def crea_utente(data: UserCreate, db: Session = Depends(get_db),
                current_user=Depends(require_admin)):
    if db.query(models.User).filter(models.User.email == data.email).first():
        raise HTTPException(400, "Email già registrata")
    u = models.User(email=data.email, nome=data.nome, ruolo=data.ruolo,
                    hashed_password=hash_password(data.password))
    db.add(u)
    db.commit()
    return {"ok": True, "id": u.id}

@app.delete("/api/utenti/{user_id}")
def elimina_utente(user_id: int, db: Session = Depends(get_db),
                   current_user=Depends(require_admin)):
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(404, "Utente non trovato")
    if u.email == current_user.email:
        raise HTTPException(400, "Non puoi eliminare te stesso")
    db.delete(u)
    db.commit()
    return {"ok": True}

# ── LOG ─────────────────────────────────────────────────────────
@app.get("/api/logs")
def get_logs(limit: int = 100, db: Session = Depends(get_db),
             current_user=Depends(get_current_user)):
    logs = db.query(models.Log).order_by(models.Log.creato_il.desc()).limit(limit).all()
    return [{"id": l.id, "operazione": l.operazione, "dettaglio": l.dettaglio,
             "livello": l.livello, "utente": l.utente_email,
             "data": l.creato_il.isoformat() if l.creato_il else ""} for l in logs]

# ── STATISTICHE CATEGORIE ───────────────────────────────────────
@app.get("/api/statistiche/categorie")
def stat_categorie(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    from sqlalchemy import func
    risultati = db.query(
        models.Prodotto.categoria1,
        func.count(models.Prodotto.id).label("totale")
    ).group_by(models.Prodotto.categoria1).order_by(func.count(models.Prodotto.id).desc()).all()
    return [{"categoria": r.categoria1 or "Non classificato", "totale": r.totale}
            for r in risultati]

# ── PING PUBBLICO (diagnostica senza auth) ──────────────────────
@app.get("/api/ping")
def ping(db: Session = Depends(get_db)):
    totale = db.query(models.Prodotto).count()
    utenti = db.query(models.User).count()
    return {"status": "ok", "prodotti_nel_db": totale, "utenti": utenti}

# ── ROOT → FRONTEND ─────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def root():
    index_path = os.path.join(static_dir, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return HTMLResponse("<h1>BindManager Pro — Frontend non trovato</h1>")
